# -*- coding: utf-8 -*-
from pandas import read_csv,read_excel
from pandas import DataFrame
from openpyxl import load_workbook
import numpy as np
def checkexcel(excelresult):
    df = read_excel("{}.xlsx".format(excelresult))
    rows = len(df)
    columns = len(df.columns)
    some_data = DataFrame.from_records(df)
    some_data = np.array(some_data).tolist()
    return rows,columns,some_data
def getcsv():
    df=read_csv('信息库.csv')
    all_data = DataFrame.from_records(df)
    return all_data
def openxl(targetexcel,someone):
    datasource=getcsv()
    someonedata=datasource[datasource['姓名'].isin([someone])]
    wb = load_workbook('{}.xlsx'.format(targetexcel))
    ws = wb.active
    listone=[]
    for item in ws.values:
        item=list(item)
        listone.append(item)
    for o in range(len(listone)):
        for p in range(len(listone[0])-1):
            if listone[o][p]!=None and listone[o][p+1]==None:
                target=listone[o][p]
                column=chr(p+66)
                listone[o][p + 1]=str(column)+str(o+1)
                try:
                    ws[listone[o][p + 1]] = list(someonedata[target])[0]
                except:
                    ws[listone[o][p + 1]] = '数据库无此项数据'
    wb.save('./{}.xlsx'.format(someone))